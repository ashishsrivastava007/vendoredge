"""
Real document extraction -- pure deterministic parsing, zero AI calls
involved. The extracted text is then fed into the EXACT SAME text
extraction pipeline already built and tested tonight for pasted text. No
new AI pathway, no new evidence-extraction logic -- just more ways to get
real text in front of the classifier that's already proven to work on
messy input.

Honesty note on scope: .xlsx, .pdf (text-based only, not scanned images),
and .eml are supported here. Scanned/image PDFs needing OCR are explicitly
NOT included -- that's a genuinely different technical undertaking (a paid
cloud OCR service or a heavy system-level dependency), not a safe addition
to make unilaterally without a real infrastructure decision first.
"""
import io
import email
import zipfile
from email import policy
from openpyxl import load_workbook
from pypdf import PdfReader


class FileExtractionError(Exception):
    """Raised for any real, expected failure -- corrupted file, wrong
    format, empty content -- so the caller can show a clear, honest
    message instead of a raw stack trace."""
    pass


def _validate_zip_container(file_bytes: bytes, *, max_entries: int = 2000, max_total_uncompressed: int = 30_000_000, max_member_uncompressed: int = 10_000_000) -> None:
    """Bound nested ZIP expansion before handing a workbook to openpyxl."""
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as archive:
            infos = [i for i in archive.infolist() if not i.is_dir()]
            if len(infos) > max_entries:
                raise FileExtractionError("This spreadsheet contains too many internal files.")
            total = 0
            for info in infos:
                size = int(info.file_size or 0)
                if size > max_member_uncompressed:
                    raise FileExtractionError("This spreadsheet contains an oversized internal file.")
                total += size
                if total > max_total_uncompressed:
                    raise FileExtractionError("This spreadsheet expands beyond the supported safety limit.")
    except zipfile.BadZipFile as e:
        raise FileExtractionError("This spreadsheet is not a valid XLSX archive.") from e


def extract_text_from_xlsx(file_bytes: bytes, max_chars: int = 8000) -> str:
    """
    Reads every cell of every sheet, row by row, and returns readable text.
    Deliberately simple -- no attempt to infer headers, merge cells, or
    guess structure. The downstream classifier is already proven to
    extract structured evidence from messy real-world text; this just gets
    the spreadsheet's actual content in front of it as plain text.

    max_chars caps the output to keep the eventual reasoning prompt a
    reasonable size -- a genuinely huge spreadsheet gets truncated with an
    honest note, not silently cut off without saying so.
    """
    _validate_zip_container(file_bytes)
    try:
        workbook = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as e:
        raise FileExtractionError(
            f"Could not read this file as a valid .xlsx spreadsheet: {type(e).__name__}"
        )

    lines = []
    for sheet in workbook.worksheets:
        lines.append(f"--- Sheet: {sheet.title} ---")
        row_count = 0
        for row in sheet.iter_rows(values_only=True):
            # Skip fully empty rows -- common in real spreadsheets, and
            # including them just adds noise without any real content.
            if all(cell is None for cell in row):
                continue
            row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
            lines.append(row_text)
            row_count += 1
            if row_count > 500:
                lines.append("[... additional rows truncated ...]")
                break

    if len(lines) <= 1:
        raise FileExtractionError("This spreadsheet appears to be empty.")

    full_text = "\n".join(lines)
    return _truncate(full_text, max_chars)


def extract_text_from_pdf(file_bytes: bytes, max_chars: int = 8000) -> str:
    """
    Extracts real, selectable text from a text-based PDF -- e.g. a
    digitally-created quote or contract, not a scanned photo of one. A
    scanned/image PDF will correctly extract as empty or near-empty text,
    since there's no OCR step here; that failure is surfaced honestly
    (FileExtractionError) rather than silently returning nothing.
    """
    try:
        reader = PdfReader(io.BytesIO(file_bytes))
    except Exception as e:
        raise FileExtractionError(
            f"Could not read this file as a valid PDF: {type(e).__name__}"
        )

    if reader.is_encrypted:
        raise FileExtractionError(
            "This PDF is password-protected. Please remove the password and try again."
        )

    pages_text = []
    for i, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        if text.strip():
            pages_text.append(f"--- Page {i + 1} ---\n{text.strip()}")

    if not pages_text:
        raise FileExtractionError(
            "No selectable text found in this PDF -- it may be a scanned image rather "
            "than a text-based document. Scanned PDFs aren't supported yet; please copy "
            "and paste the relevant details directly instead."
        )

    full_text = "\n\n".join(pages_text)
    return _truncate(full_text, max_chars)


def extract_text_from_eml(file_bytes: bytes, max_chars: int = 8000) -> str:
    """
    Extracts the subject, sender, and body from a real .eml email file
    (the standard format when someone exports/saves an email from most
    mail clients). Uses Python's built-in email library -- no new
    dependency required for this one.
    """
    try:
        msg = email.message_from_bytes(file_bytes, policy=policy.default)
    except Exception as e:
        raise FileExtractionError(f"Could not read this file as a valid email: {type(e).__name__}")

    subject = msg.get("subject", "(no subject)")
    sender = msg.get("from", "(unknown sender)")
    date = msg.get("date", "(unknown date)")

    body = ""
    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_type() == "text/plain":
                body = part.get_content()
                break
    else:
        body = msg.get_content()

    if not body or not body.strip():
        raise FileExtractionError("This email appears to have no readable text content.")

    full_text = f"From: {sender}\nDate: {date}\nSubject: {subject}\n\n{body.strip()}"
    return _truncate(full_text, max_chars)


def extract_text_from_zip(file_bytes: bytes, max_chars: int = 8000) -> str:
    """Safely extract supported text from a ZIP bundle.

    ZIPs are useful for messy procurement evidence, but a tiny compressed
    archive can expand to hundreds of MB (zip bomb) or contain thousands of
    entries. The outer upload limit alone is therefore not a sufficient
    safety boundary. We enforce entry-count, per-entry and cumulative
    uncompressed-size limits *before* reading each member.
    """
    MAX_ENTRIES = 50
    MAX_ENTRY_UNCOMPRESSED_BYTES = 5_000_000
    MAX_TOTAL_UNCOMPRESSED_BYTES = 20_000_000

    try:
        archive = zipfile.ZipFile(io.BytesIO(file_bytes))
    except zipfile.BadZipFile:
        raise FileExtractionError("Could not read this file as a valid .zip archive.")

    infos = [info for info in archive.infolist()
             if not info.is_dir() and "__MACOSX" not in info.filename
             and not info.filename.split("/")[-1].startswith(".")]
    if len(infos) > MAX_ENTRIES:
        raise FileExtractionError(
            f"This zip contains too many files ({len(infos)}). Maximum supported is {MAX_ENTRIES}."
        )

    extractors_by_ext = {
        ".xlsx": extract_text_from_xlsx,
        ".pdf": extract_text_from_pdf,
        ".eml": extract_text_from_eml,
    }

    sections, skipped = [], []
    total_uncompressed = 0
    for info in infos:
        name = info.filename
        # ZIP metadata is attacker-controlled; reject suspiciously large
        # members before decompression, rather than trusting archive.read().
        declared_size = int(info.file_size or 0)
        if declared_size > MAX_ENTRY_UNCOMPRESSED_BYTES:
            skipped.append(f"{name} (member exceeds 5MB uncompressed limit)")
            continue
        if total_uncompressed + declared_size > MAX_TOTAL_UNCOMPRESSED_BYTES:
            skipped.append(f"{name} (archive uncompressed-size budget exceeded)")
            continue

        lower_name = name.lower()
        try:
            inner_bytes = archive.read(info)
        except Exception:
            skipped.append(f"{name} (could not read from archive)")
            continue
        total_uncompressed += len(inner_bytes)
        if len(inner_bytes) > MAX_ENTRY_UNCOMPRESSED_BYTES:
            skipped.append(f"{name} (member exceeded 5MB after decompression)")
            continue

        if lower_name.endswith((".txt", ".csv")):
            try:
                text = inner_bytes.decode("utf-8", errors="replace")
                if text.strip():
                    sections.append(f"--- File: {name} ---\n{text.strip()}")
                else:
                    skipped.append(f"{name} (empty text file)")
            except Exception:
                skipped.append(f"{name} (could not decode as text)")
            continue

        matched_ext = next((ext for ext in extractors_by_ext if lower_name.endswith(ext)), None)
        if not matched_ext:
            skipped.append(f"{name} (unsupported format)")
            continue

        try:
            extracted = extractors_by_ext[matched_ext](inner_bytes, max_chars=max_chars)
            sections.append(f"--- File: {name} ---\n{extracted}")
        except FileExtractionError as e:
            skipped.append(f"{name} ({e})")

    if not sections:
        detail = "; ".join(skipped) if skipped else "the archive appears to be empty"
        raise FileExtractionError(f"No readable content found in this zip file ({detail}).")

    full_text = "\n\n".join(sections)
    if skipped:
        full_text += "\n\n[Note: could not read the following from this zip: " + "; ".join(skipped[:20]) + "]"
    return _truncate(full_text, max_chars)

def _truncate(text: str, max_chars: int) -> str:
    if len(text) > max_chars:
        return text[:max_chars] + "\n[... content truncated, file was larger than expected ...]"
    return text
